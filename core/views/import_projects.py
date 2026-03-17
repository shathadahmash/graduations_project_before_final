import openpyxl
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.views.decorators.csrf import csrf_exempt

from core.models import (
    User, Project, ProjectState, University, 
    College, Department, City, Branch, Role, Staff, Program, Student, AcademicAffiliation
)

# ==========================================================
# Header Mapping (Maintained exactly as your current order)
# ==========================================================
AR_HEADER_MAP = {
    "(عربي)عنوان المشروع": "title",
    "عنوان المشروع (انجليزي)": "title_en",
    "الملخص": "description",
    "اسم المشرف": "supervisor_first_name",        
    "اسم المشرف المشارك": "co_first_name",
    "سنة بداية المشروع": "start_year", 
    "سنة نهاية المشروع": "end_year",
    "المجال": "field",
    "الأدوات": "tools",
    "نوع المشروع": "project_type",
    "الجامعة": "university",
    "الكلية": "college",
    "القسم": "department",
    "البرنامج": "program",
    "المحافظة": "city",
    "أسماء الطلاب": "students_names",
    "أرقام قيد الطلاب": "students_ids",
    "ارقام هواتف الطلاب": "students_phones",
}

PROJECT_TYPE_MAP = {
    "حكومي": "Governmental",
    "شركات خارجية": "External",
    "مقترح": "Proposed",
}

REQUIRED_KEYS = ["title", "description", "start_year", "university", "college", "department"]

# ==========================================================
# Helpers
# ==========================================================
def _str(v): return "" if v is None else str(v).strip()
def _normalize(v): return " ".join(_str(v).split())
def _to_int(v):
    try: return int(float(_str(v)))
    except: return None

def read_excel_projects(file_obj):
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
        header_row = 2
        columns = {}
        for col in range(1, ws.max_column + 1):
            cell_val = _str(ws.cell(row=header_row, column=col).value)
            if cell_val in AR_HEADER_MAP:
                columns[col] = AR_HEADER_MAP[cell_val]

        found_keys = set(columns.values())
        missing = [k for k in REQUIRED_KEYS if k not in found_keys]
        if missing:
            return [], [{"row": 2, "message": f"الحقول التالية مفقودة: {', '.join(missing)}"}]

        rows_data = []
        for row_idx in range(3, ws.max_row + 1):
            row_dict = {}
            has_data = False
            for col_idx, key in columns.items():
                val = ws.cell(row=row_idx, column=col_idx).value
                row_dict[key] = val
                if val is not None: has_data = True
            if has_data: rows_data.append((row_idx, row_dict))
        return rows_data, []
    except Exception as e:
        return [], [{"row": 0, "message": f"فشل قراءة الملف: {str(e)}"}]

# ==========================================================
# Commit Import Logic
# ==========================================================
@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def import_projects_commit(request):
    f = request.FILES.get("file")
    if not f: return Response({"detail": "لم يتم رفع ملف"}, status=400)

    rows, file_errors = read_excel_projects(f)
    if file_errors: return Response({"errors": file_errors}, status=400)

    created_projects = 0
    updated_projects = 0
    created_users = 0

    # Ensure base objects exist
    supervisor_role, _ = Role.objects.get_or_create(type="Supervisor", defaults={"role_type": "Faculty"})
    state_obj, _ = ProjectState.objects.get_or_create(name="Accepted")

    # 1. Helper Function Defined inside the view
    def create_import_user(full_name, default_username, phone=None):
        name_parts = full_name.split()
        first_name = name_parts[0] if len(name_parts) > 0 else ""
        last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
        
        if len(name_parts) > 1:
            generated_username = f"{name_parts[0]}_{name_parts[-1]}"
        else:
            generated_username = default_username

        user_obj, created = User.objects.get_or_create(
            username=generated_username,
            defaults={
                "first_name": first_name,
                "last_name": last_name,
                "name": "", 
                "phone": phone
            }
        )
        if created:
            user_obj.set_password("123456")
            user_obj.save()
        return user_obj, created

    # 2. The Main Loop (All logic must be inside here)
    with transaction.atomic():
        for excel_row, row in rows:
            # --- A. Hierarchy Setup ---
            uni_name = _normalize(row.get("university"))
            if not uni_name: continue
            
            uni, _ = University.objects.get_or_create(uname_ar=uni_name)
            city_name = _normalize(row.get("city")) or "صنعاء"
            city_obj, _ = City.objects.get_or_create(bname_ar=city_name)
            branch_obj, _ = Branch.objects.get_or_create(university=uni, city=city_obj)
            col, _ = College.objects.get_or_create(branch=branch_obj, name_ar=_normalize(row.get("college")))
            dep, _ = Department.objects.get_or_create(college=col, name=_normalize(row.get("department")))
            
            prog_name = _normalize(row.get("program"))
            prog_obj = None
            if prog_name:
                prog_obj, _ = Program.objects.get_or_create(department=dep, p_name=prog_name)

            # --- B. Project Creation ---
            raw_type = _normalize(row.get("project_type"))
            db_project_type = PROJECT_TYPE_MAP.get(raw_type, 'Proposed')

            p, project_created = Project.objects.update_or_create(
                title=_normalize(row["title"]),
                defaults={
                    "title_en": _normalize(row.get("title_en")),
                    "project_type": db_project_type,
                    "description": _str(row.get("description")),
                    "start_date": _to_int(row.get("start_year")),
                    "end_date": _to_int(row.get("end_year")),
                    "field": _str(row.get("field")),
                    "tools": _str(row.get("tools")),
                    "state": state_obj,
                    "university": uni,
                    "branch": branch_obj,
                    "college": col,
                    "department": dep,
                    "program": prog_obj,
                }
            )

            # --- C. Staff/Supervisor Processing (INSIDE LOOP) ---
            sup_full_name = _normalize(row.get("supervisor_first_name"))
            if sup_full_name:
                backup_username = f"sup_{uni.pk}_{excel_row}"
                sup_user, u_created = create_import_user(sup_full_name, backup_username)
                
                if u_created: created_users += 1
                
                Staff.objects.get_or_create(user=sup_user, defaults={"role": supervisor_role})
                AcademicAffiliation.objects.get_or_create(
                    user=sup_user, university=uni, college=col, department=dep,
                    defaults={'start_date': timezone.now().date()}
                )
                p.created_by = sup_user 
                p.save()

            # --- D. Students Processing (INSIDE LOOP) ---
            s_names = [_normalize(x) for x in _str(row.get("students_names")).split(",") if _normalize(x)]
            s_ids = [_normalize(x) for x in _str(row.get("students_ids")).split(",") if _normalize(x)]
            s_phones = [_normalize(x) for x in _str(row.get("students_phones")).split(",") if _normalize(x)]

            for i, full_name in enumerate(s_names):
                sid = s_ids[i] if i < len(s_ids) else None 
                sphone = s_phones[i] if i < len(s_phones) else None
                
                if sid:
                    name_parts = full_name.split()
                    s_user, s_user_created = User.objects.get_or_create(
                        username=sid,
                        defaults={
                            "first_name": name_parts[0] if name_parts else "",
                            "last_name": " ".join(name_parts[1:]) if len(name_parts) > 1 else "",
                            "name": "",
                            "phone": sphone
                        }
                    )
                    if s_user_created:
                        s_user.set_password("123456")
                        s_user.save()
                else:
                    backup_sid = f"std_{p.id}_{i}"
                    s_user, s_user_created = create_import_user(full_name, backup_sid, sphone)

                if s_user_created: created_users += 1

                Student.objects.update_or_create(
                    user=s_user,
                    defaults={
                        "student_id": sid or s_user.username,
                        "phone": sphone,
                        "university": uni,
                        "college": col,
                        "department": dep,
                        "program": prog_obj,
                        "status": 'active'
                    }
                )

                AcademicAffiliation.objects.get_or_create(
                    user=s_user, university=uni, college=col, department=dep,
                    defaults={'start_date': timezone.now().date()}
                )

            if project_created: created_projects += 1
            else: updated_projects += 1

    return Response({
        "message": "تم الاستيراد بنجاح 🎉",
        "created_projects": created_projects,
        "updated_projects": updated_projects,
        "users_created": created_users
    })

# ==========================================================
# VALIDATION API
# ==========================================================

@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def import_projects_validate(request):
    f = request.FILES.get("file")

    if not f:
        return Response({"detail": "لم يتم رفع ملف"}, status=400)

    rows, file_errors = read_excel_projects(f)

    if file_errors:
        return Response({
            "errors": file_errors,
            "valid_rows": 0,
            "invalid_rows": 0
        })

    errors, valid_rows = validate_project_rows(rows)

    return Response({
        "total_rows": len(rows),
        "valid_rows": valid_rows,
        "invalid_rows": len(errors),
        "errors": errors
    })

# ==========================================================
# EXCEL TEMPLATE GENERATOR (LOGO & DOCUMENTATION REMOVED)
# ==========================================================

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def import_projects_template(request):
    # ... (Pre-fill logic remains the same) ...
    pre_city = request.GET.get("pre_city_name") or "صنعاء"
    pre_uni = request.GET.get("pre_university_name") or "جامعة صنعاء"
    pre_coll = request.GET.get("pre_college_name") or "كلية الحاسوب"
    pre_dept = request.GET.get("pre_department_name") or "علوم الحاسوب"
    pre_prog = request.GET.get("pre_program_name") or "تكنولوجيا المعلومات"

    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    ws.sheet_view.rightToLeft = True

    headers = list(AR_HEADER_MAP.keys())
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    
    # Define our Brand Color
    brand_color = "FF312583"

    # ==========================================================
    # ROW 1: TITLE (COLOR: 312583, TEXT: WHITE, HEIGHT: 40)
    # ==========================================================
    ws.merge_cells(f'A1:{last_col}1')
    title_cell = ws['A1']
    title_cell.value = "قالب استيراد بيانات المشاريع والطلاب"
    title_cell.font = Font(name='Arial', bold=True, size=16, color="FFFFFFFF")
    title_cell.fill = PatternFill(start_color=brand_color, end_color=brand_color, fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # ==========================================================
# ==========================================================
    # ROW 2: HEADERS (TRANSPARENT BACKGROUND, BRAND TEXT)
    # ==========================================================
    ws.row_dimensions[2].height = 30 
    
    # Change text color to the brand color (312583) since background is now white
    header_font = Font(name='Arial', bold=True, size=12, color="312583")
    
    # Set fill_type to None for transparency
    header_fill = PatternFill(fill_type=None) 
    
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Use the brand color for borders so the "transparent" cells have structure
    header_border = Border(
        left=Side(style='thin', color='312583'), 
        right=Side(style='thin', color='312583'), 
        top=Side(style='thin', color='312583'), 
        bottom=Side(style='thin', color='312583')
    )

    for col, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header_text)
        cell.font = header_font
        cell.fill = header_fill # This is now transparent
        cell.alignment = header_alignment
        cell.border = header_border
        ws.column_dimensions[get_column_letter(col)].width = 25

    headers = list(AR_HEADER_MAP.keys()) # This will now include the longer names automatically

    # ROW 3: EXAMPLE DATA
    example_data = [
            "نظام إدارة المكتبة",             # title
            "Library Management System",      # title_en
            "مشروع لإدارة الكتب والطلاب",      # description
            "محمد أحمد",                   # supervisor_first_name
            "خالد علي",                    # co_first_name
            2025,                             # start_year
            2026,                             # end_year
            "الذكاء الاصطناعي",               # field
            "Python, Django, React",          # tools
            "مقترح",                          # project_type
            pre_uni,                          # university
            pre_coll,                         # college
            pre_dept,                         # department
            pre_prog,                         # program
            pre_city,                         # city
            "أحمد علي, خالد محمد",            # students_names
            "2020101, 2020102",               # students_ids
            "771234567, 770000000"            # students_phones
    
        ]

    # Standard border for data cells
    data_border = Border(
        left=Side(style='thin', color='312583'), 
        right=Side(style='thin', color='312583'), 
        top=Side(style='thin', color='312583'), 
        bottom=Side(style='thin', color='312583')
    )

    for col, value in enumerate(example_data, start=1):
        cell = ws.cell(row=3, column=col, value=value)
        cell.alignment = Alignment(horizontal='right')
        cell.border = data_border

    # ==========================================================
    # GENERATE RESPONSE
    # ==========================================================
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="projects_template.xlsx"'
    return response
def validate_project_rows(rows):
    """
    Checks each row for data integrity without saving to the database.
    Returns (errors_list, count_of_valid_rows).
    """
    errors = []
    valid_count = 0

    for excel_row, row in rows:
        row_errors = []
        
        # 1. Check Required Fields
        for key in REQUIRED_KEYS:
            if not _normalize(row.get(key)):
                # We use the Arabic name from our map for the error message
                ar_name = [k for k, v in AR_HEADER_MAP.items() if v == key][0]
                row_errors.append({
                    "row": excel_row,
                    "field": ar_name,
                    "message": f"الحقل '{ar_name}' مطلوب ولا يمكن أن يكون فارغاً"
                })

        # 2. Validate Year Logic
        start_year = _to_int(row.get("start_year"))
        end_year = _to_int(row.get("end_year"))
        
        if start_year and (start_year < 2000 or start_year > 2100):
            row_errors.append({
                "row": excel_row,
                "field": "سنة بداية المشروع",
                "message": "سنة البداية غير منطقية"
            })
            
        if start_year and end_year and end_year < start_year:
            row_errors.append({
                "row": excel_row,
                "field": "سنة نهاية المشروع",
                "message": "سنة النهاية لا يمكن أن تكون قبل سنة البداية"
            })

        # 3. Check Students Data Consistency
        # Ensure if names are provided, IDs are also provided (standard practice)
        s_names = [_normalize(x) for x in _str(row.get("students_names")).split(",") if _normalize(x)]
        s_ids = [_normalize(x) for x in _str(row.get("students_ids")).split(",") if _normalize(x)]
        
        if len(s_names) > 0 and len(s_ids) == 0:
            row_errors.append({
                "row": excel_row,
                "field": "أرقام قيد الطلاب",
                "message": "يجب إدخال أرقام القيد طالما تم إدخال أسماء الطلاب"
            })

        if row_errors:
            errors.extend(row_errors)
        else:
            valid_count += 1

    return errors, valid_count

