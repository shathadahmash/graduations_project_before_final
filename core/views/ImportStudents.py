# core/views/import_students.py
from io import BytesIO
import re
import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response

from core.models import (
    User,
    Student,
    AcademicAffiliation,
    UserRoles,
    Role,
    University,
    College,
    Department,
    Program,
)

# ---------------------------
# Arabic header mapping
# ---------------------------
AR_STUDENT_HEADER_MAP = {
    "الرقم الوطني": "CID",
    "الاسم الأول": "first_name",
    "الاسم الأخير": "last_name",
    "البريد الإلكتروني": "email",
    "الهاتف": "phone",
    "الجنس": "gender",
    "رقم الطالب": "student_id",
    "الحالة": "status",
    "الجامعة": "university",
    "الكلية": "college",
    "القسم": "department",
    "البرنامج": "program",
    "سنة التسجيل": "enrollment_year",
    "سنة التخرج": "graduation_year",   # NEW
}

# Required Arabic header labels (must appear in row 2)
REQUIRED_HEADERS_AR = ["الرقم الوطني", "الاسم الأول", "الاسم الأخير", "رقم الطالب"]

VALID_GENDERS = {"ذكر", "انثى"}
VALID_STATUS = {"نشط", "موقوف", "متخرج", "منسحب"}

# ---------------------------
# Helpers
# ---------------------------
def _str(v):
    return "" if v is None else str(v).strip()

def _normalize(v):
    return " ".join(_str(v).split())

def _is_cid_valid(cid):
    s = _str(cid)
    return s.isdigit() and len(s) == 12

def _parse_enrollment_year(year_str):
    if not year_str:
        return None
    s = _str(year_str)
    m = re.match(r"(\d{4})\D*(\d{4})?", s)
    if not m:
        return None
    try:
        start_year = int(m.group(1))
        return datetime.date(start_year, 9, 1)
    except:
        return None

def _make_unique_username(base):
    uname = base.lower().strip()
    uname = re.sub(r"\s+", "_", uname)
    # This regex now allows: 
    # a-z (English), 0-9 (Numbers), _ (Underscore), 
    # and \u0600-\u06FF (The entire Arabic Unicode block)
    uname = re.sub(r"[^a-z0-9_\u0600-\u06FF]", "", uname)
    
    if not User.objects.filter(username=uname).exists():
        return uname
    
    counter = 1
    while True:
        candidate = f"{uname}{counter}"
        if not User.objects.filter(username=candidate).exists():
            return candidate
        counter += 1
def read_excel_students(file_obj):
    """
    Reads the Excel file, identifies headers from row 2, 
    and returns a list of (row_number, row_dict) tuples.
    """
    rows = []
    file_errors = []
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active

        # Row 2 contains the Arabic headers
        header_row_idx = 2
        headers_ar = [cell.value for cell in ws[header_row_idx]]

        # Validate that required headers exist
        for req in REQUIRED_HEADERS_AR:
            if req not in headers_ar:
                file_errors.append({
                    "row": header_row_idx, 
                    "field": "headers", 
                    "message": f"العمود المطلوب '{req}' غير موجود"
                })

        if file_errors:
            return [], file_errors

        # Process data starting from row 3
        for row_idx, row_cells in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            # Skip completely empty rows
            if not any(row_cells):
                continue

            # Create a dictionary mapping the internal English keys to the cell values
            row_data = {}
            for col_idx, cell_value in enumerate(row_cells):
                if col_idx < len(headers_ar):
                    ar_header = headers_ar[col_idx]
                    eng_key = AR_STUDENT_HEADER_MAP.get(ar_header)
                    if eng_key:
                        row_data[eng_key] = cell_value
            
            rows.append((row_idx, row_data))

    except Exception as e:
        file_errors.append({"row": 0, "field": "file", "message": f"تعذر قراءة ملف Excel: {str(e)}"})

    return rows, file_errors
# ---------------------------
# Excel reader for students
# ---------------------------
def validate_student_rows(rows, preselected=None):
    errors = []
    valid_rows = 0
    seen_cids = set()
    seen_student_ids = set()

    for excel_row, row in rows:
        row_errors = []
        cid = _str(row.get("CID"))
        first_name = _str(row.get("first_name"))
        last_name = _str(row.get("last_name"))
        student_id = _str(row.get("student_id"))
        gender = _str(row.get("gender"))
        status = _str(row.get("status"))
        graduation_year = _str(row.get("graduation_year"))  # <-- defined here

        # Required checks
        if not cid:
            row_errors.append(("الرقم الوطني", "الرقم الوطني مطلوب"))
        else:
            if not _is_cid_valid(cid):
                row_errors.append(("الرقم الوطني", "الرقم الوطني يجب أن يكون 12 رقماً"))
            if cid in seen_cids:
                row_errors.append(("الرقم الوطني", "تم تكرار الرقم الوطني في الملف"))
            seen_cids.add(cid)

        if not first_name:
            row_errors.append(("الاسم الأول", "الاسم الأول مطلوب"))

        if not last_name:
            row_errors.append(("الاسم الأخير", "الاسم الأخير مطلوب"))

        if not student_id:
            row_errors.append(("رقم الطالب", "رقم الطالب مطلوب"))
        else:
            if student_id in seen_student_ids:
                row_errors.append(("رقم الطالب", "تم تكرار رقم الطالب في الملف"))
            seen_student_ids.add(student_id)

        # Gender validation
        if gender and gender not in VALID_GENDERS:
            row_errors.append(("الجنس", "القيمة يجب أن تكون ذكر أو انثى"))

        # Status validation
        if status and status not in VALID_STATUS:
            row_errors.append(("الحالة", "الحالة غير معروفة"))

        # Graduation year validation
        if graduation_year and not graduation_year.isdigit():
            row_errors.append(("سنة التخرج", "سنة التخرج يجب أن تكون رقماً"))

        if row_errors:
            for field, message in row_errors:
                errors.append({"row": excel_row, "field": field, "message": message})
        else:
            valid_rows += 1

    return errors, valid_rows


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def import_students_template(request):
    """
    Logic is UNCHANGED. Design is updated to match Project Import style.
    """
    uni = request.GET.get("pre_university_id") or request.GET.get("pre_university_name")
    coll = request.GET.get("pre_college_id") or request.GET.get("pre_college_name")
    dept = request.GET.get("pre_department_id") or request.GET.get("pre_department_name")
    prog = request.GET.get("pre_program_id") or request.GET.get("pre_program_name")
    year = request.GET.get("pre_enrollment_year")

    include_cols = [
        "الرقم الوطني",
        "الاسم الأول",
        "الاسم الأخير",
        "البريد الإلكتروني",
        "الهاتف",
        "الجنس",
        "رقم الطالب",
        "الحالة",
    ]
    if not uni: include_cols.append("الجامعة")
    if not coll: include_cols.append("الكلية")
    if not dept: include_cols.append("القسم")
    if not prog: include_cols.append("البرنامج")
    if not year: include_cols.append("سنة التسجيل ")
    
    include_cols.append("سنة التخرج")

    wb = Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True
    ws.title = "استيراد الطلاب"

    # Style Constants
    brand_color_hex = "312583"
    brand_fill = PatternFill(start_color=f"FF{brand_color_hex}", end_color=f"FF{brand_color_hex}", fill_type="solid")
    white_text = Font(name='Arial', bold=True, size=16, color="FFFFFFFF")
    
    # Header Style (Transparent + Brand Blue Text)
    header_font = Font(name='Arial', bold=True, size=12, color=brand_color_hex)
    header_border = Border(
        left=Side(style='thin', color=brand_color_hex), 
        right=Side(style='thin', color=brand_color_hex), 
        top=Side(style='thin', color=brand_color_hex), 
        bottom=Side(style='thin', color=brand_color_hex)
    )

    # ==========================================================
    # ROW 1: TITLE (MATCHING DESIGN)
    # ==========================================================
    num_cols = len(include_cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value="قالب استيراد بيانات الطلاب")
    title_cell.font = white_text
    title_cell.fill = brand_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # ==========================================================
    # ROW 2: HEADERS (TRANSPARENT + BRAND BORDER)
    # ==========================================================
    ws.row_dimensions[2].height = 30
    for col, header in enumerate(include_cols, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(fill_type=None)  # Transparent
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border
        ws.column_dimensions[get_column_letter(col)].width = 25

    # ==========================================================
    # ROW 3: EXAMPLE ROW (LOGIC UNCHANGED)
    # ==========================================================
    example = [
        "123456789012", "فاطمة", "الحسن", "fatima@example.com", "777777777",
        "انثى", "STU001", "نشط"
    ]
    if "الجامعة" in include_cols: example.append("جامعة صنعاء")
    if "الكلية" in include_cols: example.append("كلية الهندسة")
    if "القسم" in include_cols: example.append("قسم الحاسوب")
    if "البرنامج" in include_cols: example.append("بكالوريوس حاسوب")
    if "السنة الدراسية" in include_cols: example.append("2021-2022")
    if "سنة التخرج" in include_cols: example.append("2025")

    data_border = Border(
        left=Side(style='thin', color='D9D9D9'), 
        right=Side(style='thin', color='D9D9D9'), 
        top=Side(style='thin', color='D9D9D9'), 
        bottom=Side(style='thin', color='D9D9D9')
    )

    for col, value in enumerate(example, start=1):
        cell = ws.cell(row=3, column=col, value=value)
        cell.alignment = Alignment(horizontal='right')
        cell.border = data_border

    # ==========================================================
    # RESPONSE GENERATION
    # ==========================================================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="students_import_template.xlsx"'
    return response

# ---------------------------
# Validation API
# ---------------------------
@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def import_students_validate(request):
    f = request.FILES.get("file")
    if not f:
        return Response({"detail": "لم يتم رفع ملف"}, status=400)

    # Build preselected dict
    preselected = {}
    for key, mapped in [
        ("pre_university_id", "university"),
        ("pre_university_name", "university"),
        ("pre_college_id", "college"),
        ("pre_college_name", "college"),
        ("pre_department_id", "department"),
        ("pre_department_name", "department"),
        ("pre_program_id", "program"),
        ("pre_program_name", "program"),
        ("pre_enrollment_year", "enrollment_year"),
    ]:
        val = request.data.get(key)
        if val:
            preselected[mapped] = val

    rows, file_errors = read_excel_students(f)
    if file_errors:
        return Response({"errors": file_errors, "valid_rows": 0, "invalid_rows": 0})

    # Validate rows and extract errors
    errors, _ = validate_student_rows(rows, preselected=preselected)
    
    # Calculate how many already exist vs new
    created_count = 0
    updated_count = 0
    for _, row in rows:
        cid = _str(row.get("CID"))
        # Check if user exists by CID
        if cid and User.objects.filter(CID=cid).exists():
            updated_count += 1
        else:
            created_count += 1

# 1. Get the actual count of valid rows from the validation function
    errors, valid_rows_count = validate_student_rows(rows, preselected=preselected)

    # 2. Return the real numbers regardless of whether errors exist elsewhere
    return Response({
        "total_rows": len(rows),
        "valid_rows": valid_rows_count, # Use the actual count from the validator
        "created_count": created_count, # Send the actual count
        "updated_count": updated_count, # Send the actual count
        "invalid_rows": len(errors),
        "errors": errors
    })

# ---------------------------
# Commit API
# ---------------------------
@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def import_students_commit(request):
    f = request.FILES.get("file")
    if not f:
        return Response({"detail": "لم يتم رفع ملف"}, status=400)

    # Build preselected objects or name tuples
    preselected = {}
    # University
    uni_id = request.data.get("pre_university_id")
    uni_name = request.data.get("pre_university_name")
    if uni_id:
        try:
            preselected["university"] = University.objects.get(pk=int(uni_id))
        except:
            return Response({"detail": "معرف الجامعة المسبق غير صالح"}, status=400)
    elif uni_name:
        preselected["university"] = ("name", _normalize(uni_name))

    # College
    coll_id = request.data.get("pre_college_id")
    coll_name = request.data.get("pre_college_name")
    if coll_id:
        try:
            preselected["college"] = College.objects.get(pk=int(coll_id))
        except:
            return Response({"detail": "معرف الكلية المسبق غير صالح"}, status=400)
    elif coll_name:
        preselected["college"] = ("name", _normalize(coll_name))

    # Department
    dept_id = request.data.get("pre_department_id")
    dept_name = request.data.get("pre_department_name")
    if dept_id:
        try:
            preselected["department"] = Department.objects.get(pk=int(dept_id))
        except:
            return Response({"detail": "معرف القسم المسبق غير صالح"}, status=400)
    elif dept_name:
        preselected["department"] = ("name", _normalize(dept_name))

    # Program
    prog_id = request.data.get("pre_program_id")
    prog_name = request.data.get("pre_program_name")
    if prog_id:
        try:
            preselected["program"] = Program.objects.get(pk=int(prog_id))
        except:
            return Response({"detail": "معرف البرنامج المسبق غير صالح"}, status=400)
    elif prog_name:
        preselected["program"] = ("name", _normalize(prog_name))

    # Enrollment year
    pre_year = request.data.get("pre_enrollment_year")
    if pre_year:
        preselected["enrollment_year"] = _parse_enrollment_year(pre_year)

    rows, file_errors = read_excel_students(f)
    if file_errors:
        return Response({"errors": file_errors}, status=400)

    errors, valid_rows = validate_student_rows(rows, preselected=preselected)
    if errors:
        return Response({
            "message": "لا يمكن الاستيراد بسبب وجود أخطاء",
            "errors": errors
        }, status=400)

    created_users = 0
    updated_users = 0
    created_students = 0
    updated_students = 0
    created_affiliations = 0
    assigned_roles = 0
    row_errors = []

    student_role, _ = Role.objects.get_or_create(type="Student", defaults={"role_type": "Student"})

    with transaction.atomic():
        for excel_row, row in rows:
            try:
                cid = _str(row.get("CID"))
                first_name = _str(row.get("first_name"))
                last_name = _str(row.get("last_name"))
                email = _str(row.get("email"))
                phone = _str(row.get("phone"))
                gender = _normalize(row.get("gender"))
                student_id = _str(row.get("student_id"))
                status = _str(row.get("status")) or "نشط"
                enrollment_year_val = _str(row.get("enrollment_year"))

                # Resolve affiliation objects
                uni_obj = None
                if isinstance(preselected.get("university"), University):
                    uni_obj = preselected.get("university")
                elif isinstance(preselected.get("university"), tuple):
                    uni_obj, _ = University.objects.get_or_create(uname_ar=preselected["university"][1])
                else:
                    uni_name = _normalize(row.get("university"))
                    if uni_name:
                        uni_obj, _ = University.objects.get_or_create(uname_ar=uni_name)

                coll_obj = None
                if isinstance(preselected.get("college"), College):
                    coll_obj = preselected.get("college")
                elif isinstance(preselected.get("college"), tuple):
                    coll_name = preselected["college"][1]
                    if uni_obj:
                        branch = uni_obj.branch_set.first()
                        coll_obj, _ = College.objects.get_or_create(branch=branch, name_ar=coll_name)
                    else:
                        coll_obj, _ = College.objects.get_or_create(name_ar=coll_name)
                else:
                    coll_name = _normalize(row.get("college"))
                    if coll_name:
                        if uni_obj:
                            coll_qs = College.objects.filter(name_ar=coll_name, branch__university=uni_obj)
                            if coll_qs.exists():
                                coll_obj = coll_qs.first()
                            else:
                                branch = uni_obj.branch_set.first()
                                coll_obj, _ = College.objects.get_or_create(branch=branch, name_ar=coll_name)
                        else:
                            coll_obj, _ = College.objects.get_or_create(name_ar=coll_name)

                dep_obj = None
                if isinstance(preselected.get("department"), Department):
                    dep_obj = preselected.get("department")
                elif isinstance(preselected.get("department"), tuple):
                    dep_name = preselected["department"][1]
                    if coll_obj:
                        dep_obj, _ = Department.objects.get_or_create(college=coll_obj, name=dep_name)
                    else:
                        dep_obj, _ = Department.objects.get_or_create(name=dep_name)
                else:
                    dep_name = _normalize(row.get("department"))
                    if dep_name:
                        if coll_obj:
                            dep_obj, _ = Department.objects.get_or_create(college=coll_obj, name=dep_name)
                        else:
                            dep_obj, _ = Department.objects.get_or_create(name=dep_name)

                prog_obj = None
                if isinstance(preselected.get("program"), Program):
                    prog_obj = preselected.get("program")
                elif isinstance(preselected.get("program"), tuple):
                    prog_name = preselected["program"][1]
                    if dep_obj:
                        prog_obj, _ = Program.objects.get_or_create(p_name=prog_name, department=dep_obj)
                    else:
                        prog_obj, _ = Program.objects.get_or_create(p_name=prog_name)
                else:
                    prog_name = _normalize(row.get("program"))
                    if prog_name:
                        if dep_obj:
                            prog_obj, _ = Program.objects.get_or_create(p_name=prog_name, department=dep_obj)
                        else:
                            prog_obj, _ = Program.objects.get_or_create(p_name=prog_name)

                # Enrollment start date
                if preselected.get("enrollment_year"):
                    start_date = preselected.get("enrollment_year")
                else:
                    start_date = _parse_enrollment_year(enrollment_year_val) or timezone.now().date()
                graduation_year_val = _str(row.get("graduation_year"))
                graduation_year_int = int(graduation_year_val) if graduation_year_val.isdigit() else None

# --- START OF FIXED MODIFIED SECTION ---

                # 1. Capture names with a case-insensitive/robust check
                # We check for the keys "first_name" and "last_name"
                raw_first = ""
                raw_last = ""
                
                for k, v in row.items():
                    if k and str(k).strip().lower() == "first_name":
                        raw_first = _str(v)
                    if k and str(k).strip().lower() == "last_name":
                        raw_last = _str(v)

                # 2. Construct the username components
                # We remove all spaces and convert to lowercase
                clean_first = raw_first.lower().replace(" ", "")
                clean_last = raw_last.lower().replace(" ", "")

                # 3. Create the base username
                if clean_first and clean_last:
                    base_username = f"{clean_first}_{clean_last}"
                elif clean_first or clean_last:
                    # If only one name exists, don't leave a hanging underscore
                    base_username = clean_first or clean_last
                else:
                    # Fallback to IDs if names are completely missing in this row
                    fallback = _str(row.get("student_id")) or _str(row.get("CID"))
                    base_username = f"user_{fallback}" if fallback else "student_user"

                # 4. Finalize unique username (handles duplicates)
                username = _make_unique_username(base_username)

                # --- Database Operations ---
                user_obj = None
                if cid:
                    user_qs = User.objects.filter(CID=cid)
                    if user_qs.exists():
                        user_obj = user_qs.first()
                        user_obj.first_name = raw_first
                        user_obj.last_name = raw_last
                        user_obj.name = ""  # Keep empty as requested
                        # Only update username if it's currently empty or a placeholder
                        if not user_obj.username or "_" in user_obj.username:
                            user_obj.username = username
                        user_obj.save()
                        updated_users += 1
                    else:
                        user_obj = User.objects.create(
                            username=username,
                            CID=cid,
                            first_name=raw_first,
                            last_name=raw_last,
                            name="", 
                            email=_str(row.get("email")) or None,
                            phone=_str(row.get("phone")) or None,
                            gender=_normalize(row.get("gender")) or None,
                        )
                        user_obj.set_unusable_password()
                        user_obj.save()
                        created_users += 1

                # --- END OF FIXED MODIFIED SECTION ---                # --- END OF MODIFIED SECTION ---

                # Student profile
                student_obj = None
                if hasattr(user_obj, "student_profile"):
                    student_obj = user_obj.student_profile
                    ...
                    student_obj.graduation_year = graduation_year_int or student_obj.graduation_year
                    student_obj.save()
                    updated_students += 1
                else:
                    student_obj = Student.objects.create(
                        user=user_obj,
                        student_id=student_id,
                        phone=phone or None,
                        status=status or "نشط",
                        university=uni_obj if isinstance(uni_obj, University) else None,
                        college=coll_obj if isinstance(coll_obj, College) else None,
                        department=dep_obj if isinstance(dep_obj, Department) else None,
                        program=prog_obj if isinstance(prog_obj, Program) else None,
                        graduation_year=graduation_year_int,   # NEW
                    )
                    created_students += 1


                # AcademicAffiliation
                if uni_obj:
                    aff, created_aff = AcademicAffiliation.objects.get_or_create(
                        user=user_obj,
                        university=uni_obj,
                        start_date=start_date,
                        defaults={
                            "college": coll_obj if isinstance(coll_obj, College) else None,
                            "department": dep_obj if isinstance(dep_obj, Department) else None,
                            "end_date": None
                        }
                    )
                    if created_aff:
                        created_affiliations += 1

                # Assign Student role
                ur, created_ur = UserRoles.objects.get_or_create(user=user_obj, role=student_role)
                if created_ur:
                    assigned_roles += 1

            except Exception as e:
                row_errors.append({"row": excel_row, "field": "عام", "message": f"خطأ أثناء المعالجة: {str(e)}"})

    result = {
        "message": "انتهى الاستيراد",
        "total_rows": len(rows),
        "valid_rows": valid_rows,
        "created_users": created_users,
        "updated_users": updated_users,
        "created_students": created_students,
        "updated_students": updated_students,
        "created_affiliations": created_affiliations,
        "assigned_roles": assigned_roles,
        "row_errors": row_errors
    }
    return Response(result)
