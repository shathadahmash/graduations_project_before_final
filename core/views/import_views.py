from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
from django.utils.dateparse import parse_date

import openpyxl

from core.models import User, Role, UserRoles, AcademicAffiliation, University


# ---------------------------
# Helpers
# ---------------------------

def is_system_manager(user) -> bool:
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return UserRoles.objects.filter(user=user, role__type__iexact="system manager").exists()


def normalize_role(raw) -> str:
    """
    Keep it role (كما طلبتي) لكن نطبع فقط عشان تقل الأخطاء من Excel: spaces/case/_/-
    """
    if raw is None:
        return ""
    x = str(raw).strip().lower()
    x = x.replace("_", "-")
    x = " ".join(x.split())
    return x


def _cell_to_str(v) -> str:
    return "" if v is None else str(v).strip()


def _parse_excel_date(v):
    # Excel may provide datetime/date object
    if v is None or str(v).strip() == "":
        return None
    if hasattr(v, "date"):
        try:
            return v.date()
        except Exception:
            pass
    # else parse string
    return parse_date(str(v).strip())


def read_excel(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    # header row
    headers = []
    for c in ws[1]:
        headers.append(_cell_to_str(c.value))

    header_map = {h: i for i, h in enumerate(headers)}

    required = ["email", "username", "first_name", "last_name", "role", "university_name_ar", "start_date"]
    missing = [c for c in required if c not in header_map]
    if missing:
        return None, [{
            "row": 1,
            "field": ",".join(missing),
            "message": "Missing required columns",
            "value": None
        }]

    rows = []
    for r in range(2, ws.max_row + 1):
        # skip empty row
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == "" for v in values):
            continue

        row = {col: ws.cell(row=r, column=header_map[col] + 1).value for col in required}
        rows.append((r, row))
    return rows, []


def validate_rows(rows):
    errors = []
    valid = []

    for excel_row, row in rows:
        email = _cell_to_str(row.get("email"))
        username = _cell_to_str(row.get("username"))
        first_name = _cell_to_str(row.get("first_name"))
        last_name = _cell_to_str(row.get("last_name"))
        role_raw = row.get("role")
        university_name_ar = _cell_to_str(row.get("university_name_ar"))
        start_date_raw = row.get("start_date")

        # required fields
        if not email:
            errors.append({"row": excel_row, "field": "email", "message": "email is required", "value": row.get("email")})
        if not username:
            errors.append({"row": excel_row, "field": "username", "message": "username is required", "value": row.get("username")})
        if not first_name:
            errors.append({"row": excel_row, "field": "first_name", "message": "first_name is required", "value": row.get("first_name")})
        if not last_name:
            errors.append({"row": excel_row, "field": "last_name", "message": "last_name is required", "value": row.get("last_name")})
        if role_raw is None or str(role_raw).strip() == "":
            errors.append({"row": excel_row, "field": "role", "message": "role is required", "value": row.get("role")})
        if not university_name_ar:
            errors.append({"row": excel_row, "field": "university_name_ar", "message": "university_name_ar is required", "value": row.get("university_name_ar")})

        # date parse
        start_date = _parse_excel_date(start_date_raw)
        if not start_date:
            errors.append({"row": excel_row, "field": "start_date", "message": "start_date must be YYYY-MM-DD", "value": start_date_raw})

        # role exists?
        role_norm = normalize_role(role_raw)
        role_obj = Role.objects.filter(type__iexact=role_norm).first()
        if role_norm and not role_obj:
            errors.append({"row": excel_row, "field": "role", "message": "role not found in Role.type", "value": role_raw})

        # university exists? (we can create later in commit, but validate to alert user)
        # If you prefer NOT to require it exists, remove this check.
        # Here: we allow creating, so no error.

        # if row has errors, don't add to valid
        row_has_errors = any(e["row"] == excel_row for e in errors)
        if not row_has_errors:
            valid.append({
                "email": email,
                "username": username,
                "first_name": first_name,
                "last_name": last_name,
                "role_norm": role_norm,
                "university_name_ar": university_name_ar,
                "start_date": start_date,
            })

    return valid, errors


# ---------------------------
# API Endpoints
# ---------------------------

@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def import_users_validate(request):
    if not is_system_manager(request.user):
        return Response({"detail": "Forbidden: System Manager only."}, status=403)

    f = request.FILES.get("file")
    if not f:
        return Response({"detail": "No file uploaded. Use form-data field name: file"}, status=400)

    rows, file_errors = read_excel(f)
    if file_errors:
        return Response({
            "total_rows": 0,
            "valid_rows": 0,
            "invalid_rows": len(file_errors),
            "errors": file_errors,
        }, status=400)

    valid, errors = validate_rows(rows)

    return Response({
        "total_rows": len(rows),
        "valid_rows": len(valid),
        "invalid_rows": len(errors),
        "errors": errors,
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def import_users_commit(request):
    if not is_system_manager(request.user):
        return Response({"detail": "Forbidden: System Manager only."}, status=403)

    f = request.FILES.get("file")
    if not f:
        return Response({"detail": "No file uploaded. Use form-data field name: file"}, status=400)

    rows, file_errors = read_excel(f)
    if file_errors:
        return Response({
            "total_rows": 0,
            "valid_rows": 0,
            "invalid_rows": len(file_errors),
            "errors": file_errors,
        }, status=400)

    valid, errors = validate_rows(rows)
    if errors:
        return Response({
            "total_rows": len(rows),
            "valid_rows": len(valid),
            "invalid_rows": len(errors),
            "errors": errors,
            "message": "Fix validation errors then retry commit."
        }, status=400)

    created_users = 0
    updated_users = 0
    roles_assigned = 0
    affiliations_created = 0

    with transaction.atomic():
        for item in valid:
            email = item["email"]
            username = item["username"]
            first_name = item["first_name"]
            last_name = item["last_name"]
            role_norm = item["role_norm"]
            uname_ar = item["university_name_ar"]
            start_date = item["start_date"]

            # 1) Create or Update User
            user_obj = User.objects.filter(email__iexact=email).first()
            if not user_obj:
                user_obj = User.objects.filter(username__iexact=username).first()

            if user_obj:
                user_obj.email = email
                user_obj.username = username
                user_obj.first_name = first_name
                user_obj.last_name = last_name
                user_obj.name = f"{first_name} {last_name}".strip()
                user_obj.save()
                updated_users += 1
            else:
                user_obj = User.objects.create(
                    username=username,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    name=f"{first_name} {last_name}".strip(),
                )
                user_obj.set_password("password123")  # اختياري
                user_obj.save()
                created_users += 1

            # 2) Assign Role (keep it role)
            role_obj = Role.objects.filter(type__iexact=role_norm).first()
            if role_obj:
                _, created = UserRoles.objects.get_or_create(user=user_obj, role=role_obj)
                if created:
                    roles_assigned += 1

            # 3) University (create if not exists)
            uni = University.objects.filter(uname_ar__iexact=uname_ar).first()
            if not uni:
                uni = University.objects.create(uname_ar=uname_ar)

            # 4) AcademicAffiliation
            # unique_together = (user, university, start_date)
            aff, created = AcademicAffiliation.objects.get_or_create(
                user=user_obj,
                university=uni,
                start_date=start_date,
                defaults={
                    "college": None,
                    "department": None,
                    "end_date": None,
                }
            )
            if created:
                affiliations_created += 1
            else:
                # تحديث إن لزم (هنا نضمن الجامعة/التاريخ ثابتين)
                aff.university = uni
                aff.save()

    return Response({
        "total_rows": len(rows),
        "valid_rows": len(valid),
        "invalid_rows": 0,
        "created_users": created_users,
        "updated_users": updated_users,
        "roles_assigned": roles_assigned,
        "affiliations_created": affiliations_created,
        "errors": [],
        "message": "Import completed successfully."
    })
